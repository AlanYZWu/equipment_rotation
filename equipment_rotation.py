import openpyxl as xl
import random
import math

# Load availability information
availability_book = xl.load_workbook(filename="Equipment Rotation Availability.xlsx")
availability_page = availability_book["Availability"]

# Load equipment rotation
rotation_book = xl.load_workbook(filename="Equipment Rotation.xlsx")
rotation_page = rotation_book["Apr 24"]

# Sets denoting who can take each type of equipment
member_set = set()
lion_set = set()
drum_set = set()
box_set = set()


def text_to_boolean(text):
    if "y" in text.lower() or "m" in text.lower():
        return True
    else:
        return False


def generate_sets():
    # Create Members and add to member_set
    for r in range(2, availability_page.max_row + 1):
        member_info = availability_page[r]
        if text_to_boolean(member_info[1].value):
            lion_set.add(member_info[0].value)
            member_set.add(member_info[0])

        if text_to_boolean(member_info[2].value):
            drum_set.add(member_info[0].value)
            member_set.add(member_info[0])

        if text_to_boolean(member_info[3].value):
            box_set.add(member_info[0].value)
            member_set.add(member_info[0])

        if text_to_boolean(member_info[4].value):
            member_set.add(member_info[0])


def check_used(row, name):
    for c in range(2, rotation_page.max_column - 2):
        if rotation_page.cell(row=row, column=c).value is not None \
                and name in rotation_page.cell(row=row, column=c).value:
            return True
        elif name == "Angie/Small" and "S" in rotation_page.cell(row=row, column=1).value:
            return True
        elif rotation_page.cell(row=row, column=c).value is not None \
                and "Lions" in rotation_page.cell(row=row, column=c).value \
                and "Lions" in name:
            return True
        elif rotation_page.cell(row=row, column=c).value is not None \
                and "Seniors" in rotation_page.cell(row=row, column=c).value\
                and "Seniors" in name:
            return True
    return False


def check_used_helper(row, name):
    for c in range(2, 14):
        if rotation_page.cell(row=row, column=c).value is not None \
                and name in rotation_page.cell(row=row, column=c).value:
            return True
        elif name == "Angie/Small" and "S" in rotation_page.cell(row=row, column=1).value:
            return True
    return False


# Dictionary of to track usage
usage = dict()
usage["-"] = 0

# Add troupe members to dictionary
for cell in availability_page["A"]:
    usage[cell.value] = 0

# Initialize dictionary values
row = 2
while rotation_page.cell(row=row, column=1).value is not None:
    for col in range(2, rotation_page.max_column - 2):
        if rotation_page.cell(row=row, column=col).value is not None:
            usage[rotation_page.cell(row=row, column=col).value] = \
                usage[rotation_page.cell(row=row, column=col).value] + 1

    row = row + 1
del usage["-"], usage["Name"]

generate_sets()
usage_limit = math.ceil(((rotation_page.max_column - 5) * (row - 2)) / (member_set.__len__()))

row = 2

while rotation_page.cell(row=row, column=1).value is not None:
    for col in range(2, rotation_page.max_column - 2):
        if rotation_page.cell(row=row, column=col).value is not None:
            continue

        member = "^"
        assigned_helper = False
        equipment_type = rotation_page.cell(row=1, column=col).value

        if "Drum" in equipment_type:
            member = random.choice(list(drum_set))
            while check_used(row, member):
                member = random.choice(list(drum_set))
        elif "Box" in equipment_type:
            member = random.choice(list(box_set))
            while check_used(row, member):
                member = random.choice(list(box_set))
        else:
            member = random.choice(list(lion_set))
            while check_used(row, member):
                member = random.choice(list(lion_set))

        rotation_page.cell(row=row, column=col, value=member)
        usage[member] = usage[member] + 1

        if usage[member] > usage_limit:
            drum_set.discard(member)
            box_set.discard(member)
            lion_set.discard(member)
        rotation_book.save('Equipment Rotation.xlsx')

    row = row + 1
